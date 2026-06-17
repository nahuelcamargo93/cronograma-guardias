import pandas as pd
import openpyxl

wb = openpyxl.load_workbook("Cronograma_Servicio_Kinesiologia_Julio26.xlsx")
sheet_names = wb.sheetnames
print("Hojas disponibles:", sheet_names)

if "Vista por Personal" in sheet_names:
    sheet = wb["Vista por Personal"]
    
    # Leer filas del excel
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
        
    # Imprimir las primeras filas y buscar las celdas que contienen FLR
    print("\n--- Vista por Personal (Primeras filas) ---")
    for r in data[:10]:
        # Filtrar valores no nulos para simplificar la vista
        valores = [str(x) for x in r if x is not None]
        print(valores[:15])
        
    print("\n--- Conteos de FLR en la hoja ---")
    conteo_flr = 0
    for r in data:
        for val in r:
            if val == "FLR":
                conteo_flr += 1
    print("Cantidad total de celdas con 'FLR':", conteo_flr)
else:
    print("La hoja 'Vista por Personal' no se encuentra.")
