import openpyxl
import pandas as pd

file_name = "Cronograma_Servicio_Kinesiologia_Julio26.xlsx"

def run():
    wb = openpyxl.load_workbook(file_name)
    print("--- Hojas en el Excel ---")
    print(wb.sheetnames)
    
    sheet_name = "Reporte Histórico"
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = list(sheet.values)
        cols = data[0]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=cols)
        print(f"\n--- Contenido de la hoja '{sheet_name}' (Flores) ---")
        print(df[df['Personal'].str.contains('Flores', na=False, case=False)])
    else:
        print(f"Error: La hoja '{sheet_name}' no existe.")
        
if __name__ == '__main__':
    run()
