import openpyxl
import pandas as pd

wb = openpyxl.load_workbook("para_importar_enfermeria.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\nSheet: {sheet_name}")
    ws = wb[sheet_name]
    
    # Print merged ranges
    merged_ranges = ws.merged_cells.ranges
    print(f"Number of merged cell ranges: {len(merged_ranges)}")
    for r in list(merged_ranges)[:10]:
        print(f"  Range: {r.coord}")
        # Let's inspect the value in the top-left cell of the merged range
        top_left_cell = ws[r.coord.split(':')[0]]
        print(f"    Top-left value: {top_left_cell.value} (Type: {type(top_left_cell.value)})")
