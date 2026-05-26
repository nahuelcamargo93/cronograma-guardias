import openpyxl
import pandas as pd

wb = openpyxl.load_workbook("para_importar_enfermeria.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    df = pd.read_excel("para_importar_enfermeria.xlsx", sheet_name=sheet_name, header=None)
    
    # We need to map row index to person name
    # Let's find start row of names
    start_row = 0
    for r in range(min(5, df.shape[0])):
        val = str(df.iloc[r, 0]).strip().upper()
        if "APELLIDO" in val or "NOMBRE" in val:
            start_row = r + 1
            break
            
    print(f"\n--- Sheet: {sheet_name} (start_row: {start_row}) ---")
    for r_range in ws.merged_cells.ranges:
        top_left_cell = ws[r_range.coord.split(':')[0]]
        val = top_left_cell.value
        
        # Get row index (1-indexed in excel, convert to 0-indexed)
        min_row = r_range.min_row - 1
        max_row = r_range.max_row - 1
        min_col = r_range.min_col - 1
        max_col = r_range.max_col - 1
        
        if min_row >= start_row:
            name = df.iloc[min_row, 0]
            print(f"  Range: {r_range.coord} | Row name: '{name}' | Val: {val} | Cols: {min_col} to {max_col}")
