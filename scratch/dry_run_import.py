import sqlite3
import openpyxl
import pandas as pd
from datetime import datetime, date, timedelta

excel_path = "para_importar_enfermeria.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Load DB names for validation
conn = sqlite3.connect("cronograma_inteligente.db")
db_names = set(row[0] for row in conn.execute("SELECT nombre FROM personal WHERE servicio_id = 2").fetchall())
conn.close()

MONTH_MAP = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3,
    "ABRIL": 4, "MAYO": 5, "JUNIO": 6
}

# Shift map configuration
VALID_SHIFTS = {"M", "T", "TN", "N", "TNN", "MT"}

def map_name(name):
    if not name or pd.isna(name):
        return None
    name = str(name).strip().upper()
    if not name or "GRUPO" in name or name == "NAN":
        return None
    if name in db_names:
        return name
    if "CARRERAS" in name:
        return "CARRERAS FLAVIA"
    if "ABELENDA" in name:
        return "ABELENDA GRISELL"
    if "ALCARAZ FRANCISCO" in name or name == "ALCARAZ FRANCISCO":
        return "ALCARAZ FRANCISO"
    if "BARROSO ERIKA" in name or name == "BARROSO ERIKA":
        return "BARROSO ERICA"
    if "BORIA MAIRA" in name or name == "BORIA MAIRA":
        return "BORIA MAYRA"
    if "CALDERON" in name:
        return "CALDERON MARIA JOSE"
    if "GOMEZ LOURDEZ" in name or name == "GOMEZ LOURDEZ":
        return "GOMEZ LOURDES"
    if "GRABOVIECKI" in name:
        return "GRABOVIECKI FERNANDA"
    if "PEREIRA" in name:
        return "PEREIRA CRISTINA"
    if name in ["ESCUDERO YANET", "LUCERO SABRINA"]:
        return name
    return name

# We will collect everything here
all_simulated_guardias = []
all_simulated_licencias = []
unhandled_cells = []
all_parsed_names = set()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    
    # 1. Detect month
    month_num = None
    for k, m in MONTH_MAP.items():
        if k in sheet_name.upper():
            month_num = m
            break
    if not month_num:
        print(f"Skipping sheet {sheet_name} (unable to detect month)")
        continue
        
    print(f"\nProcessing sheet: {sheet_name} (Month: {month_num})")
    
    # 2. Identify start row
    header_row_idx = None
    for r in range(min(5, df.shape[0])):
        val = str(df.iloc[r, 0]).strip().upper()
        if "APELLIDO" in val or "NOMBRE" in val:
            header_row_idx = r
            break
            
    if header_row_idx is None:
        print(f"Error: could not find header row in sheet {sheet_name}")
        continue
        
    # 3. Map columns to dates
    col_to_date = {}
    row_header = df.iloc[header_row_idx]
    for c in range(1, len(row_header)):
        val = row_header[c]
        if pd.isna(val):
            continue
        # Check if it is a date
        if isinstance(val, (datetime, date)):
            col_to_date[c] = val.date() if isinstance(val, datetime) else val
        else:
            try:
                # Try parsing as integer day of month
                day = int(float(val))
                if 1 <= day <= 31:
                    col_to_date[c] = date(2026, month_num, day)
            except ValueError:
                # Not a day column
                pass
                
    print(f"  Mapped {len(col_to_date)} date columns. Range: {min(col_to_date.values())} to {max(col_to_date.values())}")
    
    # 4. Parse merged cell ranges for licencias
    # We will build a mapping of (row_idx, col_idx) -> license info
    # Note: openpyxl rows and cols are 1-based, we will map them to 0-based
    cell_license_map = {}
    for r_range in ws.merged_cells.ranges:
        # Check if range is within dates header
        min_row = r_range.min_row - 1
        max_row = r_range.max_row - 1
        min_col = r_range.min_col - 1
        max_col = r_range.max_col - 1
        
        if min_row <= header_row_idx:
            # Header or title merged cells, ignore
            continue
            
        top_left_val = ws.cell(row=r_range.min_row, column=r_range.min_col).value
        # Check if any columns are date columns
        range_dates = [col_to_date[c] for c in range(min_col, max_col + 1) if c in col_to_date]
        if not range_dates:
            continue
            
        # Determine license type
        val_str = str(top_left_val).strip().upper() if top_left_val else ""
        lic_type = None
        if "LAR" in val_str:
            lic_type = "LAR"
        elif "LPP" in val_str or "COMPENSACION" in val_str or "COMP." in val_str or "CM" in val_str or "C" == val_str or "LIC" in val_str or not top_left_val:
            lic_type = "LPP"
            
        if lic_type:
            # Associate this license with all cells in the range
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell_license_map[(r, c)] = {
                        "tipo": lic_type,
                        "desc": val_str if val_str else "CELDA VACIA UNIFICADA",
                        "start": min(range_dates),
                        "end": max(range_dates)
                    }
                    
    # Let's count parsed merged licenses
    # Each range is unique. Let's collect them
    unique_merged_lics = set()
    for (r, c), info in cell_license_map.items():
        unique_merged_lics.add((r, info["tipo"], info["start"], info["end"]))
    print(f"  Parsed {len(unique_merged_lics)} merged cell licenses.")

    # 5. Parse rows
    for r in range(header_row_idx + 1, df.shape[0]):
        raw_name = df.iloc[r, 0]
        mapped = map_name(raw_name)
        if not mapped:
            continue
            
        all_parsed_names.add(mapped)
        
        # Iterate over date columns
        for c, dt in col_to_date.items():
            # Check if this cell is covered by a license
            if (r, c) in cell_license_map:
                lic_info = cell_license_map[(r, c)]
                # Add to simulated licencias (we'll deduplicate later)
                all_simulated_licencias.append({
                    "nombre": mapped,
                    "tipo": lic_info["tipo"],
                    "fecha_inicio": lic_info["start"].isoformat(),
                    "fecha_fin": lic_info["end"].isoformat(),
                    "origen": f"Merged range val: {lic_info['desc']}"
                })
                continue
                
            # Otherwise, read cell value
            cell_val = df.iloc[r, c]
            if pd.isna(cell_val):
                continue
                
            cell_str = str(cell_val).strip()
            if not cell_str:
                continue
                
            # Strip spaces and uppercase
            cell_clean = cell_str.upper()
            
            # Check if it is a valid shift
            if cell_clean in VALID_SHIFTS:
                all_simulated_guardias.append({
                    "nombre": mapped,
                    "fecha": dt.isoformat(),
                    "turno": cell_clean,
                    "horas": 12 if cell_clean in ["TNN", "MT"] else 6
                })
            elif cell_clean == "TTN":
                all_simulated_guardias.append({
                    "nombre": mapped,
                    "fecha": dt.isoformat(),
                    "turno": "TNN", # Map TTN to TNN
                    "horas": 12
                })
            elif cell_clean in ["FF", "FS"]:
                # Franco, skip
                pass
            elif "LAR" in cell_clean:
                # Individual cell LAR license
                all_simulated_licencias.append({
                    "nombre": mapped,
                    "tipo": "LAR",
                    "fecha_inicio": dt.isoformat(),
                    "fecha_fin": dt.isoformat(),
                    "origen": f"Cell value: {cell_clean}"
                })
            elif cell_clean in ["COMPENSACION HORARIA", "COMPENSACION H", "COMP. HS", "CM", "C"]:
                # Individual cell license (not merged)
                all_simulated_licencias.append({
                    "nombre": mapped,
                    "tipo": "LPP",
                    "fecha_inicio": dt.isoformat(),
                    "fecha_fin": dt.isoformat(),
                    "origen": f"Cell value: {cell_clean}"
                })
            else:
                unhandled_cells.append({
                    "sheet": sheet_name,
                    "row": r,
                    "col": c,
                    "name": mapped,
                    "date": dt.isoformat(),
                    "val": cell_val
                })

# Print summaries
print("\n======================================")
print("IMPORT SIMULATION SUMMARY")
print("======================================")
print(f"Total mapped personnel: {len(all_parsed_names)}")
new_people = all_parsed_names - db_names
print(f"New personnel to be registered ({len(new_people)}): {list(new_people)}")

print(f"\nTotal Simulated Guardias: {len(all_simulated_guardias)}")
print(f"Total Raw Simulated Licencias: {len(all_simulated_licencias)}")

# Deduplicate licencias
# Licencias might overlap or be repeated because of cell-by-cell collection.
# For merged licenses, we can group by (nombre, tipo, fecha_inicio, fecha_fin) and deduplicate.
df_lics = pd.DataFrame(all_simulated_licencias)
if not df_lics.empty:
    df_lics_uniq = df_lics.drop_duplicates(subset=["nombre", "tipo", "fecha_inicio", "fecha_fin"])
    print(f"Total Unique Simulated Licencias: {len(df_lics_uniq)}")
    print("\nSample of Unique Licencias (first 10):")
    print(df_lics_uniq.head(10).to_string())
    print("\nLAR Count:", len(df_lics_uniq[df_lics_uniq['tipo'] == 'LAR']))
    print("LPP Count:", len(df_lics_uniq[df_lics_uniq['tipo'] == 'LPP']))

if unhandled_cells:
    print(f"\nWARNING: {len(unhandled_cells)} unhandled cell values found!")
    for idx, uc in enumerate(unhandled_cells[:15]):
        print(f"  {idx+1:02d}. Sheet: {uc['sheet']} | Name: {uc['name']} | Date: {uc['date']} | Value: {repr(uc['val'])}")
else:
    print("\nSUCCESS: No unhandled cells found!")
