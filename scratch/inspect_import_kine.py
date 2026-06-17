import pandas as pd
import openpyxl

file_name = "para_importar_kineJunio.xlsx"

def run():
    wb = openpyxl.load_workbook(file_name)
    print("--- Hojas en el Excel para importar ---")
    print(wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Filas de la hoja: {sheet_name} ---")
        try:
            df = pd.read_excel(file_name, sheet_name=sheet_name)
            print("Columnas:", df.columns.tolist())
            print("Primeras 5 filas:")
            print(df.head(5))
        except Exception as e:
            print(f"Error al leer la hoja {sheet_name}: {e}")

if __name__ == '__main__':
    run()
