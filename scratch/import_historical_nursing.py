import sqlite3
import openpyxl
import pandas as pd
from datetime import datetime, date, timedelta

excel_path = "para_importar_enfermeria.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Load DB names for validation
conn = sqlite3.connect("cronograma_inteligente.db")
db_names = set(row[0] for row in conn.execute("SELECT nombre FROM personal WHERE servicio_id = 2").fetchall())

MONTH_MAP = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3,
    "ABRIL": 4, "MAYO": 5, "JUNIO": 6
}

VALID_SHIFTS = {"M", "T", "TN", "N", "TNN", "MT"}

# Argentina National Holidays 2026
ARGENTINA_HOLIDAYS_2026 = {
    date(2026, 1, 1),   # Año Nuevo
    date(2026, 2, 16),  # Carnaval
    date(2026, 2, 17),  # Carnaval
    date(2026, 3, 23),  # Feriado puente
    date(2026, 3, 24),  # Día de la Memoria
    date(2026, 4, 2),   # Día de Malvinas
    date(2026, 4, 3),   # Viernes Santo
    date(2026, 5, 1),   # Día del Trabajo
    date(2026, 5, 25),  # Revolución de Mayo
    date(2026, 6, 15),  # Paso a la Inmortalidad de Güemes
    date(2026, 6, 20)   # Paso a la Inmortalidad de Belgrano
}

def map_name(name):
    if not name or pd.isna(name):
        return None
    name = str(name).strip().upper()
    if not name or "GRUPO" in name or name == "NAN":
        return None
    if name in db_names:
        return name
    if "CARRERAS" in name:
        return "CARRERAS FLAVIA"
    if "ABELENDA" in name:
        return "ABELENDA GRISELL"
    if "ALCARAZ FRANCISCO" in name or name == "ALCARAZ FRANCISCO":
        return "ALCARAZ FRANCISO"
    if "BARROSO ERIKA" in name or name == "BARROSO ERIKA":
        return "BARROSO ERICA"
    if "BORIA MAIRA" in name or name == "BORIA MAIRA":
        return "BORIA MAYRA"
    if "CALDERON" in name:
        return "CALDERON MARIA JOSE"
    if "GOMEZ LOURDEZ" in name or name == "GOMEZ LOURDEZ":
        return "GOMEZ LOURDES"
    if "GRABOVIECKI" in name:
        return "GRABOVIECKI FERNANDA"
    if "PEREIRA" in name:
        return "PEREIRA CRISTINA"
    if name in ["ESCUDERO YANET", "LUCERO SABRINA"]:
        return name
    return name

all_simulated_guardias = []
all_simulated_licencias = []
unhandled_cells = []
all_parsed_names = set()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
    
    # 1. Detect month
    month_num = None
    for k, m in MONTH_MAP.items():
        if k in sheet_name.upper():
            month_num = m
            break
    if not month_num:
        continue
        
    # 2. Identify start row
    header_row_idx = None
    for r in range(min(5, df.shape[0])):
        val = str(df.iloc[r, 0]).strip().upper()
        if "APELLIDO" in val or "NOMBRE" in val:
            header_row_idx = r
            break
            
    if header_row_idx is None:
        continue
        
    # 3. Map columns to dates
    col_to_date = {}
    row_header = df.iloc[header_row_idx]
    for c in range(1, len(row_header)):
        val = row_header[c]
        if pd.isna(val):
            continue
        if isinstance(val, (datetime, date)):
            col_to_date[c] = val.date() if isinstance(val, datetime) else val
        else:
            try:
                day = int(float(val))
                if 1 <= day <= 31:
                    col_to_date[c] = date(2026, month_num, day)
            except ValueError:
                pass
                
    # 4. Parse merged cell ranges for licencias
    cell_license_map = {}
    for r_range in ws.merged_cells.ranges:
        min_row = r_range.min_row - 1
        max_row = r_range.max_row - 1
        min_col = r_range.min_col - 1
        max_col = r_range.max_col - 1
        
        if min_row <= header_row_idx:
            continue
            
        top_left_val = ws.cell(row=r_range.min_row, column=r_range.min_col).value
        range_dates = [col_to_date[c] for c in range(min_col, max_col + 1) if c in col_to_date]
        if not range_dates:
            continue
            
        val_str = str(top_left_val).strip().upper() if top_left_val else ""
        lic_type = None
        if "LAR" in val_str:
            lic_type = "LAR"
        elif "CM" in val_str or val_str == "C":
            lic_type = "CM"
        elif "LPP" in val_str or "COMPENSACION" in val_str or "COMP." in val_str or "LIC" in val_str or not top_left_val:
            lic_type = "LPP"
            
        if lic_type:
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell_license_map[(r, c)] = {
                        "tipo": lic_type,
                        "desc": val_str if val_str else "CELDA VACIA UNIFICADA",
                        "start": min(range_dates),
                        "end": max(range_dates)
                    }

    # 5. Parse rows
    for r in range(header_row_idx + 1, df.shape[0]):
        raw_name = df.iloc[r, 0]
        mapped = map_name(raw_name)
        if not mapped:
            continue
            
        all_parsed_names.add(mapped)
        
        for c, dt in col_to_date.items():
            if (r, c) in cell_license_map:
                lic_info = cell_license_map[(r, c)]
                all_simulated_licencias.append({
                    "nombre": mapped,
                    "tipo": lic_info["tipo"],
                    "fecha_inicio": lic_info["start"].isoformat(),
                    "fecha_fin": lic_info["end"].isoformat()
                })
                continue
                
            cell_val = df.iloc[r, c]
            if pd.isna(cell_val):
                continue
                
            cell_str = str(cell_val).strip()
            if not cell_str:
                continue
                
            cell_clean = cell_str.upper()
            
            if cell_clean in VALID_SHIFTS:
                all_simulated_guardias.append({
                    "nombre": mapped,
                    "fecha": dt.isoformat(),
                    "turno": cell_clean,
                    "horas": 12 if cell_clean in ["TNN", "MT"] else 6
                })
            elif cell_clean == "TTN":
                all_simulated_guardias.append({
                    "nombre": mapped,
                    "fecha": dt.isoformat(),
                    "turno": "TNN",
                    "horas": 12
                })
            elif cell_clean in ["FF", "FS"]:
                pass
            elif "LAR" in cell_clean:
                all_simulated_licencias.append({
                    "nombre": mapped,
                    "tipo": "LAR",
                    "fecha_inicio": dt.isoformat(),
                    "fecha_fin": dt.isoformat()
                })
            elif "CM" in cell_clean or cell_clean == "C":
                all_simulated_licencias.append({
                    "nombre": mapped,
                    "tipo": "CM",
                    "fecha_inicio": dt.isoformat(),
                    "fecha_fin": dt.isoformat()
                })
            elif cell_clean in ["COMPENSACION HORARIA", "COMPENSACION H", "COMP. HS", "COMPENSACION"]:
                all_simulated_licencias.append({
                    "nombre": mapped,
                    "tipo": "LPP",
                    "fecha_inicio": dt.isoformat(),
                    "fecha_fin": dt.isoformat()
                })
            else:
                unhandled_cells.append((sheet_name, r, c, mapped, dt, cell_val))

if unhandled_cells:
    print(f"ERROR: Found {len(unhandled_cells)} unhandled cells:")
    for uc in unhandled_cells:
        print(uc)
    exit(1)

# Group and merge Excel licencias
def merge_intervals(intervals):
    if not intervals:
        return []
    intervals = sorted(intervals, key=lambda x: x[0])
    merged = [list(intervals[0])]
    for current in intervals[1:]:
        prev = merged[-1]
        if current[0] <= prev[1] + timedelta(days=1):
            prev[1] = max(prev[1], current[1])
        else:
            merged.append(list(current))
    return [tuple(x) for x in merged]

excel_lic_intervals = {}
for lic in all_simulated_licencias:
    key = (lic["nombre"], lic["tipo"])
    fi = date.fromisoformat(lic["fecha_inicio"])
    ff = date.fromisoformat(lic["fecha_fin"])
    excel_lic_intervals.setdefault(key, []).append((fi, ff))

merged_excel_lics = []
for (nombre, tipo), intervals in excel_lic_intervals.items():
    merged = merge_intervals(intervals)
    for fi, ff in merged:
        merged_excel_lics.append({
            "nombre": nombre,
            "tipo": tipo,
            "fecha_inicio": fi.isoformat(),
            "fecha_fin": ff.isoformat()
        })

print(f"Parsed {len(all_simulated_guardias)} guardias and merged {len(merged_excel_lics)} licenses from Excel.")

# Let's perform DB changes
conn.execute("BEGIN TRANSACTION")
try:
    # 1. Register missing personnel
    for p in ["ESCUDERO YANET", "LUCERO SABRINA"]:
        conn.execute("INSERT OR IGNORE INTO personal (nombre, rol, servicio_id, categoria) VALUES (?, 'Rotativo', 2, '')", (p,))
        print(f"Registered/checked personnel: {p}")

    # 2. Inactivate TNN for service 2
    conn.execute("UPDATE turnos_config SET activo = 0 WHERE servicio_id = 2 AND nombre = 'TNN'")
    print("Inactivated shift TNN in turnos_config")

    # 3. Clean up overlapping draft cronogramas for service 2 in Jan-Jun 2026
    c_ids = [row[0] for row in conn.execute("""
        SELECT DISTINCT c.id
        FROM cronogramas c
        JOIN guardias g ON g.cronograma_id = c.id
        JOIN personal p ON g.nombre = p.nombre
        WHERE p.servicio_id = 2 AND c.fecha_inicio >= '2026-01-01' AND c.fecha_fin <= '2026-06-30'
    """).fetchall()]
    if c_ids:
        print(f"Deleting existing cronogramas for service 2 in Jan-Jun 2026: {c_ids}")
        placeholders = ",".join("?" for _ in c_ids)
        conn.execute(f"DELETE FROM guardias WHERE cronograma_id IN ({placeholders})", c_ids)
        conn.execute(f"DELETE FROM bloques_finde_largo WHERE cronograma_id IN ({placeholders})", c_ids)
        conn.execute(f"DELETE FROM flr_asignados WHERE cronograma_id IN ({placeholders})", c_ids)
        conn.execute(f"DELETE FROM semanas_categorias WHERE cronograma_id IN ({placeholders})", c_ids)
        conn.execute(f"DELETE FROM cronogramas WHERE id IN ({placeholders})", c_ids)

    # 4. Create the 6 new cronogramas
    month_to_cronograma_id = {}
    monthly_ranges = {
        1: ("2026-01-01", "2026-01-31"),
        2: ("2026-02-01", "2026-02-28"),
        3: ("2026-03-01", "2026-03-31"),
        4: ("2026-04-01", "2026-04-30"),
        5: ("2026-05-01", "2026-05-31"),
        6: ("2026-06-01", "2026-06-30")
    }
    for m, (start_dt, end_dt) in monthly_ranges.items():
        cur = conn.execute("""
            INSERT INTO cronogramas (fecha_inicio, fecha_fin, creado_en, notas, estado)
            VALUES (?, ?, ?, 'Historial de Enfermeria importado', 'aprobado')
        """, (start_dt, end_dt, datetime.now().isoformat(timespec='seconds')))
        month_to_cronograma_id[m] = cur.lastrowid
        print(f"Created cronograma ID {month_to_cronograma_id[m]} for month {m}")

    # 5. Insert guardias
    def is_weekend_or_holiday(dt):
        return dt.weekday() >= 5 or dt in ARGENTINA_HOLIDAYS_2026

    for g in all_simulated_guardias:
        dt_obj = date.fromisoformat(g["fecha"])
        m_num = dt_obj.month
        c_id = month_to_cronograma_id[m_num]
        es_finde = 1 if is_weekend_or_holiday(dt_obj) else 0
        conn.execute("""
            INSERT INTO guardias (cronograma_id, nombre, fecha, turno, horas, es_finde)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (c_id, g["nombre"], g["fecha"], g["turno"], g["horas"], es_finde))

    print(f"Inserted {len(all_simulated_guardias)} guardias successfully.")

    # 6. Merge and insert/update licenses
    db_lics = []
    cur = conn.execute("""
        SELECT l.id, l.nombre, l.tipo, l.fecha_inicio, l.fecha_fin
        FROM licencias l
        JOIN personal p ON l.nombre = p.nombre
        WHERE p.servicio_id = 2
    """)
    for row in cur.fetchall():
        db_lics.append({
            "id": row[0],
            "nombre": row[1],
            "tipo": row[2],
            "fecha_inicio": row[3],
            "fecha_fin": row[4],
            "merged": False
        })

    inserted_count = 0
    updated_count = 0

    for el in merged_excel_lics:
        inserted = False
        for dl in db_lics:
            if dl["nombre"] == el["nombre"] and dl["tipo"] == el["tipo"]:
                dl_fi = dl["fecha_inicio"]
                dl_ff = dl["fecha_fin"]
                el_fi = el["fecha_inicio"]
                el_ff = el["fecha_fin"]
                
                # Check overlap or adjacency (1 day diff max)
                if (dl_fi <= (date.fromisoformat(el_ff) + timedelta(days=1)).isoformat() and 
                    el_fi <= (date.fromisoformat(dl_ff) + timedelta(days=1)).isoformat()):
                    # Merge
                    new_fi = min(dl_fi, el_fi)
                    new_ff = max(dl_ff, el_ff)
                    dl["fecha_inicio"] = new_fi
                    dl["fecha_fin"] = new_ff
                    dl["merged"] = True
                    inserted = True
                    break
        if not inserted:
            conn.execute("""
                INSERT INTO licencias (nombre, tipo, fecha_inicio, fecha_fin)
                VALUES (?, ?, ?, ?)
            """, (el["nombre"], el["tipo"], el["fecha_inicio"], el["fecha_fin"]))
            inserted_count += 1
        else:
            updated_count += 1

    # Apply updates to merged licenses
    for dl in db_lics:
        if dl["merged"]:
            conn.execute("""
                UPDATE licencias
                SET fecha_inicio = ?, fecha_fin = ?
                WHERE id = ?
            """, (dl["fecha_inicio"], dl["fecha_fin"], dl["id"]))

    print(f"Licencias processing complete. Inserted: {inserted_count}, Merged/Updated existing: {updated_count}")

    # 7. Auto-calculate and save long weekend blocks
    for m, c_id in month_to_cronograma_id.items():
        start_str, end_str = monthly_ranges[m]
        start_dt = date.fromisoformat(start_str)
        end_dt = date.fromisoformat(end_str)
        dias_totales = (end_dt - start_dt).days + 1
        offset_dia = start_dt.weekday()
        
        month_feriados = [
            (f_dt - start_dt).days 
            for f_dt in ARGENTINA_HOLIDAYS_2026 
            if start_dt <= f_dt <= end_dt
        ]
        
        es_descanso = [
            (((d + offset_dia) % 7) >= 5 or d in month_feriados)
            for d in range(dias_totales)
        ]
        bloques = []
        actual = []
        for d in range(dias_totales):
            if es_descanso[d]:
                actual.append(d)
            else:
                if len(actual) >= 3:
                    bloques.append(actual[:])
                actual = []
        if len(actual) >= 3:
            bloques.append(actual)
            
        for bloque in bloques:
            fi = (start_dt + timedelta(days=bloque[0])).isoformat()
            ff = (start_dt + timedelta(days=bloque[-1])).isoformat()
            tipo = min(len(bloque), 4)
            conn.execute("""
                INSERT INTO bloques_finde_largo (cronograma_id, fecha_inicio, fecha_fin, tipo)
                VALUES (?, ?, ?, ?)
            """, (c_id, fi, ff, tipo))
            
    conn.commit()
    print("Database transaction committed successfully!")
except Exception as e:
    conn.rollback()
    print("Error occurred, rolled back database changes:", e)
    raise e
finally:
    conn.close()
