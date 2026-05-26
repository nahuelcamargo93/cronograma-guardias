import openpyxl

wb = openpyxl.load_workbook("para_importar_enfermeria.xlsx", data_only=True)
ws = wb["ENERO 26 "]
# Let's search for "CARRERAS" in the first column
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val and "CARRERAS" in str(val).upper():
        print(f"Row {r}: {repr(val)}")
conn = ws
